import streamlit as st
from datetime import datetime
from zoneinfo import ZoneInfo
import re
import io
import csv
import html as html_lib
from typing import Optional, Tuple, Dict, List
from collections import defaultdict
import difflib

import os
import base64

import gspread
from google.oauth2.service_account import Credentials

import dropbox
from dropbox.sharing import RequestedVisibility, SharedLinkSettings
from dropbox.exceptions import ApiError, AuthError

import qrcode

# Optional libs for better export / image optimization
try:
    from PIL import Image, ImageOps
    PIL_AVAILABLE = True
except Exception:
    PIL_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False


# =========================
# CONFIG
# =========================
st.set_page_config(
    page_title="JALA • Absensi QR",
    page_icon="✅",
    layout="centered",
    initial_sidebar_state="collapsed",
)

APP_CFG = st.secrets.get("app", {})
SHEET_NAME = APP_CFG.get("sheet_name", "Absensi_Karyawan")
WORKSHEET_NAME = APP_CFG.get("worksheet_name", "Log")
DROPBOX_ROOT = APP_CFG.get("dropbox_folder", "/Absensi_Selfie")
TZ_NAME = APP_CFG.get("timezone", "Asia/Jakarta")

QR_URL = APP_CFG.get("qr_url", "")
ENABLE_TOKEN = bool(APP_CFG.get("enable_token", False))
TOKEN_SECRET = str(APP_CFG.get("token", "")).strip()

DEFAULT_SHEET_ROWS = int(APP_CFG.get("sheet_rows", 10000))

# Optimasi foto untuk HP spek rendah / internet lambat
IMG_MAX_SIDE = int(APP_CFG.get("img_max_side", 1280))
IMG_JPEG_QUALITY = int(APP_CFG.get("img_jpeg_quality", 78))

# Brand / Tema JALA (bisa override via secrets)
BRAND_NAME = str(APP_CFG.get("brand_name", "JALA")).strip() or "JALA"
BRAND_TAGLINE = str(APP_CFG.get("brand_tagline", "Jala Tech")).strip() or "Jala Tech"
BRAND_PRIMARY = str(APP_CFG.get("brand_primary", "#0B66E4")).strip() or "#0B66E4"
BRAND_ACCENT = str(APP_CFG.get("brand_accent", "#46C2FF")).strip() or "#46C2FF"
BRAND_BG = str(APP_CFG.get("brand_bg", "#F5FAFF")).strip() or "#F5FAFF"

# ✅ Logo path (default ke assets/jala.png)
LOGO_PATH = str(APP_CFG.get("logo_path", "assets/jala.png")).strip() or "assets/jala.png"

COL_TIMESTAMP = "Timestamp"
COL_NAMA = "Nama"
COL_HP = "No HP/WA"
COL_POSISI = "Posisi"
COL_LINK_SELFIE = "Bukti Selfie"
COL_DBX_PATH = "Dropbox Path"

SHEET_COLUMNS = [COL_TIMESTAMP, COL_NAMA, COL_HP, COL_POSISI, COL_LINK_SELFIE, COL_DBX_PATH]


# =========================
# UI THEME (CSS)
# =========================
def inject_brand_css():
    st.markdown(
        f"""
<style>
/* Force light (mencegah komponen jadi hitam & text tidak kebaca) */
html, body {{
  color-scheme: light !important;
}}
:root {{
  color-scheme: light !important;
  --jala-primary: {BRAND_PRIMARY};
  --jala-accent: {BRAND_ACCENT};
  --jala-bg: {BRAND_BG};
  --jala-text: #0A2540;
  --jala-muted: #516579;
  --jala-border: rgba(11, 102, 228, 0.18);
  --jala-border-strong: rgba(11, 102, 228, 0.30);
  --jala-shadow: 0 10px 30px rgba(11, 102, 228, 0.14);
  --jala-card-shadow: 0 6px 16px rgba(11, 102, 228, 0.08);
}}

[data-testid="stAppViewContainer"] {{
  color-scheme: light !important;
  background: linear-gradient(180deg,
    rgba(70,194,255,0.14) 0%,
    rgba(245,250,255,1) 18%,
    #FFFFFF 100%) !important;
}}

[data-testid="stHeader"] {{
  background: rgba(255,255,255,0) !important;
}}

#MainMenu {{ visibility: hidden; }}
footer {{ visibility: hidden; }}

[data-testid="stAppViewContainer"] .main .block-container {{
  padding-top: 0.9rem;
  padding-bottom: 1.2rem;
  max-width: 720px;
}}

h1,h2,h3,h4, p, label, .stMarkdown, .stCaption {{
  color: var(--jala-text) !important;
}}

.jala-topbar {{
  background: linear-gradient(135deg, var(--jala-accent) 0%, var(--jala-primary) 62%, #0B4CC7 100%);
  border-radius: 18px;
  padding: 14px 14px;
  box-shadow: var(--jala-shadow);
  margin: 0.25rem 0 1rem 0;
  overflow: hidden;
  position: relative;
}}
.jala-topbar:before {{
  content: "";
  position: absolute;
  top: -90px;
  right: -130px;
  width: 280px;
  height: 280px;
  background: rgba(255,255,255,0.16);
  border-radius: 999px;
}}

.jala-brand {{
  display: flex;
  align-items: center;
  justify-content: center; /* center header */
  position: relative;
  z-index: 2;
  gap: 12px;
}}

.jala-brand-center {{
  display: flex;
  flex-direction: column;
  align-items: center;
  gap: 4px;
}}

.jala-logo {{
  height: 44px;
  width: auto;
  object-fit: contain;
  /* bikin logo jadi putih (logo biru -> putih) */
  filter: brightness(0) invert(1);
}}

.jala-subtitle {{
  font-size: 12px;
  color: rgba(255,255,255,0.92) !important;
  text-align: center;
  margin-top: 0;
}}

.jala-chip {{
  display: inline-flex;
  align-items: center;
  justify-content: center;
  padding: 7px 12px;
  border-radius: 999px;
  font-size: 12px;
  color: #FFFFFF !important;
  background: rgba(255,255,255,0.16);
  border: 1px solid rgba(255,255,255,0.22);
  white-space: nowrap;

  position: absolute;
  right: 14px;
  top: 50%;
  transform: translateY(-50%);
}}

@media (max-width: 520px) {{
  .jala-chip {{ display: none; }}
  .jala-logo {{ height: 40px; }}
}}

.jala-card {{
  background: rgba(255,255,255,0.96);
  border: 1px solid var(--jala-border);
  border-radius: 16px;
  padding: 14px 14px;
  box-shadow: var(--jala-card-shadow);
}}
.jala-muted {{
  color: var(--jala-muted) !important;
  font-size: 13px;
}}
.jala-divider {{
  height: 1px;
  background: rgba(11,102,228,0.12);
  margin: 12px 0;
}}

/* Inputs */
input, textarea {{
  background: #FFFFFF !important;
  color: var(--jala-text) !important;
}}
div[data-testid="stTextInput"] input,
div[data-testid="stTextArea"] textarea {{
  border-radius: 12px !important;
}}

/* Buttons */
div[data-testid="stButton"] button,
div[data-testid="stDownloadButton"] button,
div[data-testid="stFormSubmitButton"] button {{
  border-radius: 14px !important;
  padding: 0.7rem 1rem !important;
  background: #FFFFFF !important;
  color: var(--jala-text) !important;
  border: 1px solid var(--jala-border-strong) !important;
}}
div[data-testid="stButton"] button:hover,
div[data-testid="stDownloadButton"] button:hover {{
  background: rgba(11,102,228,0.06) !important;
  border-color: rgba(11,102,228,0.45) !important;
}}
div[data-testid="stButton"] button:disabled,
div[data-testid="stDownloadButton"] button:disabled,
div[data-testid="stFormSubmitButton"] button:disabled {{
  opacity: 0.7 !important;
  background: #F2F6FF !important;
  color: rgba(10,37,64,0.75) !important;
  border-color: rgba(11,102,228,0.20) !important;
}}
/* Submit gradient */
div[data-testid="stFormSubmitButton"] button {{
  background: linear-gradient(135deg, var(--jala-accent) 0%, var(--jala-primary) 82%) !important;
  color: #FFFFFF !important;
  border: 0 !important;
}}
div[data-testid="stFormSubmitButton"] button:hover {{
  filter: brightness(1.03) !important;
}}

/* File uploader */
div[data-testid="stFileUploaderDropzone"] {{
  background: #FFFFFF !important;
  border: 1px dashed rgba(11,102,228,0.45) !important;
  border-radius: 14px !important;
  padding: 12px !important;
}}
div[data-testid="stFileUploaderDropzone"] * {{
  color: var(--jala-text) !important;
}}
div[data-testid="stFileUploaderDropzone"] button {{
  background: #FFFFFF !important;
  color: var(--jala-text) !important;
  border: 1px solid rgba(11,102,228,0.30) !important;
  border-radius: 12px !important;
}}
div[data-testid="stFileUploaderDropzone"] button:hover {{
  background: rgba(11,102,228,0.06) !important;
}}

/* Expander - FIX open state yang suka berubah jadi hitam */
details[data-testid="stExpander"] > summary,
div[data-testid="stExpander"] summary,
.streamlit-expanderHeader {{
  background: #FFFFFF !important;
  color: var(--jala-text) !important;
  border: 1px solid var(--jala-border) !important;
  border-radius: 14px !important;
  padding: 10px 12px !important;
}}
details[data-testid="stExpander"][open] > summary,
div[data-testid="stExpander"] details[open] > summary {{
  background: #FFFFFF !important;
  color: var(--jala-text) !important;
}}
details[data-testid="stExpander"] > summary:hover,
div[data-testid="stExpander"] summary:hover {{
  background: rgba(11,102,228,0.06) !important;
}}
details[data-testid="stExpander"] > div,
div[data-testid="stExpander"] details > div {{
  border: 1px solid var(--jala-border) !important;
  border-top: 0 !important;
  border-radius: 0 0 14px 14px !important;
  background: rgba(255,255,255,0.96) !important;
}}
details[data-testid="stExpander"] summary * ,
div[data-testid="stExpander"] summary * ,
.streamlit-expanderHeader * {{
  color: inherit !important;
}}

/* Simple table (untuk rekap) - ringan & konsisten */
.jala-table-wrap {{
  width: 100%;
  overflow-x: auto;
  background: #FFFFFF;
  border: 1px solid var(--jala-border);
  border-radius: 14px;
  box-shadow: var(--jala-card-shadow);
}}
table.jala-table {{
  width: 100%;
  border-collapse: separate;
  border-spacing: 0;
  min-width: 520px;
}}
table.jala-table th {{
  background: rgba(245,250,255,1);
  color: var(--jala-text);
  text-align: left;
  font-weight: 700;
  padding: 10px 12px;
  border-bottom: 1px solid rgba(11,102,228,0.14);
  position: sticky;
  top: 0;
}}
table.jala-table td {{
  color: var(--jala-text);
  padding: 10px 12px;
  border-bottom: 1px solid rgba(11,102,228,0.10);
  vertical-align: top;
}}
table.jala-table tr:last-child td {{
  border-bottom: 0;
}}
</style>
        """,
        unsafe_allow_html=True,
    )


# ✅ helper logo (diletakkan sebelum render_header agar bisa dipakai)
def _abs_path(rel_path: str) -> str:
    base = os.path.dirname(__file__)
    return os.path.join(base, rel_path)


@st.cache_data(show_spinner=False)
def load_logo_data_uri(path: str) -> str:
    if not path:
        return ""
    full = _abs_path(path)
    if not os.path.exists(full):
        return ""

    ext = os.path.splitext(full)[1].lower().replace(".", "")
    mime = "png" if ext == "png" else ("webp" if ext == "webp" else "jpeg")

    with open(full, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("utf-8")
    return f"data:image/{mime};base64,{b64}"


def render_header(chip_text: str, subtitle: str):
    logo_uri = load_logo_data_uri(LOGO_PATH)
    logo_html = f'<img class="jala-logo" src="{logo_uri}" alt="{BRAND_NAME} logo"/>' if logo_uri else ""

    st.markdown(
        f"""
<div class="jala-topbar">
  <div class="jala-brand">
    <div class="jala-brand-center">
      {logo_html}
      <div class="jala-subtitle">{subtitle}</div>
    </div>
    <div class="jala-chip">{chip_text}</div>
  </div>
</div>
        """,
        unsafe_allow_html=True,
    )


inject_brand_css()


# =========================
# HELPERS
# =========================
def get_mode() -> str:
    try:
        return str(st.query_params.get("mode", "")).strip().lower()
    except Exception:
        qp = st.experimental_get_query_params()
        return (qp.get("mode", [""])[0] or "").strip().lower()


def get_token_from_url() -> str:
    try:
        return str(st.query_params.get("token", "")).strip()
    except Exception:
        qp = st.experimental_get_query_params()
        return (qp.get("token", [""])[0] or "").strip()


def sanitize_name(text: str) -> str:
    text = str(text).strip()
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[^A-Za-z0-9 _.-]", "", text)
    return text.strip()


def sanitize_phone(text: str) -> str:
    text = str(text).strip()
    if text.startswith("+"):
        return "+" + re.sub(r"\D", "", text[1:])
    return re.sub(r"\D", "", text)


def now_local():
    return datetime.now(tz=ZoneInfo(TZ_NAME))


@st.cache_data(show_spinner=False)
def build_qr_png(url: str) -> bytes:
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=8,
        border=2,
    )
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


def make_hyperlink(url: str, label: str = "Bukti Foto") -> str:
    if not url or url == "-":
        return "-"
    safe = url.replace('"', '""')
    return f'=HYPERLINK("{safe}", "{label}")'


def detect_ext_and_mime(mime: str) -> str:
    mime = (mime or "").lower()
    if "png" in mime:
        return ".png"
    return ".jpg"


def get_selfie_bytes(selfie_cam, selfie_upload) -> Tuple[Optional[bytes], str]:
    if selfie_cam is not None:
        mime = getattr(selfie_cam, "type", "") or ""
        return selfie_cam.getvalue(), detect_ext_and_mime(mime)
    if selfie_upload is not None:
        mime = getattr(selfie_upload, "type", "") or ""
        return selfie_upload.getvalue(), detect_ext_and_mime(mime)
    return None, ".jpg"


def optimize_image_bytes(img_bytes: bytes, ext: str) -> Tuple[bytes, str]:
    if not PIL_AVAILABLE:
        return img_bytes, ext
    try:
        img = Image.open(io.BytesIO(img_bytes))
        img = ImageOps.exif_transpose(img)
        if img.mode not in ("RGB", "L"):
            bg = Image.new("RGB", img.size, (255, 255, 255))
            if img.mode in ("RGBA", "LA"):
                bg.paste(img, mask=img.split()[-1])
            else:
                bg.paste(img)
            img = bg
        else:
            img = img.convert("RGB")

        w, h = img.size
        max_side = max(w, h)
        if max_side > IMG_MAX_SIDE:
            scale = IMG_MAX_SIDE / float(max_side)
            new_size = (max(1, int(w * scale)), max(1, int(h * scale)))
            img = img.resize(new_size, Image.LANCZOS)

        out = io.BytesIO()
        img.save(out, format="JPEG", quality=IMG_JPEG_QUALITY, optimize=True, progressive=True)
        return out.getvalue(), ".jpg"
    except Exception:
        return img_bytes, ext


def escape(s: str) -> str:
    return html_lib.escape(str(s if s is not None else ""))


def render_table(rows: List[Dict], columns: List[str], min_width_px: int = 520):
    """Render tabel HTML ringan (lebih cepat & konsisten daripada st.dataframe)."""
    if not rows:
        st.info("Tidak ada data.")
        return

    th = "".join([f"<th>{escape(c)}</th>" for c in columns])
    body_rows = []
    for r in rows:
        tds = "".join([f"<td>{escape(r.get(c, ''))}</td>" for c in columns])
        body_rows.append(f"<tr>{tds}</tr>")
    tbody = "".join(body_rows)

    st.markdown(
        f"""
<div class="jala-table-wrap" style="min-width: 100%;">
  <table class="jala-table" style="min-width:{min_width_px}px;">
    <thead><tr>{th}</tr></thead>
    <tbody>{tbody}</tbody>
  </table>
</div>
        """,
        unsafe_allow_html=True,
    )


# =========================
# GOOGLE SHEETS
# =========================
def auto_format_absensi_sheet(ws):
    try:
        sheet_id = ws.id
        row_count = ws.row_count
        col_widths = [170, 180, 150, 180, 140, 340]
        requests = []

        for i, w in enumerate(col_widths):
            requests.append({
                "updateDimensionProperties": {
                    "range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": i, "endIndex": i + 1},
                    "properties": {"pixelSize": w},
                    "fields": "pixelSize"
                }
            })

        requests.append({
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1},
                "cell": {"userEnteredFormat": {
                    "textFormat": {"bold": True},
                    "horizontalAlignment": "CENTER",
                    "verticalAlignment": "MIDDLE",
                    "backgroundColor": {"red": 0.93, "green": 0.95, "blue": 0.99},
                    "wrapStrategy": "WRAP"
                }},
                "fields": "userEnteredFormat(textFormat,horizontalAlignment,verticalAlignment,backgroundColor,wrapStrategy)"
            }
        })

        requests.append({
            "updateSheetProperties": {
                "properties": {"sheetId": sheet_id, "gridProperties": {"frozenRowCount": 1}},
                "fields": "gridProperties.frozenRowCount"
            }
        })

        requests.append({
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": row_count},
                "cell": {"userEnteredFormat": {"verticalAlignment": "MIDDLE", "wrapStrategy": "CLIP"}},
                "fields": "userEnteredFormat(verticalAlignment,wrapStrategy)"
            }
        })

        # Center Timestamp & HP
        requests.append({
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": row_count,
                          "startColumnIndex": 0, "endColumnIndex": 1},
                "cell": {"userEnteredFormat": {"horizontalAlignment": "CENTER"}},
                "fields": "userEnteredFormat(horizontalAlignment)"
            }
        })
        requests.append({
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": row_count,
                          "startColumnIndex": 2, "endColumnIndex": 3},
                "cell": {"userEnteredFormat": {"horizontalAlignment": "CENTER"}},
                "fields": "userEnteredFormat(horizontalAlignment)"
            }
        })

        # Wrap Dropbox Path
        requests.append({
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": row_count,
                          "startColumnIndex": 5, "endColumnIndex": 6},
                "cell": {"userEnteredFormat": {"wrapStrategy": "WRAP"}},
                "fields": "userEnteredFormat(wrapStrategy)"
            }
        })

        ws.spreadsheet.batch_update({"requests": requests})
    except Exception as e:
        print(f"Format Absensi Error: {e}")


@st.cache_resource
def connect_gsheet():
    if "gcp_service_account" not in st.secrets:
        raise RuntimeError("GSheet secrets tidak ditemukan: gcp_service_account")

    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    creds_dict = dict(st.secrets["gcp_service_account"])
    if "private_key" in creds_dict:
        creds_dict["private_key"] = creds_dict["private_key"].replace("\\n", "\n")

    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    gc = gspread.authorize(creds)
    return gc.open(SHEET_NAME)


def get_or_create_ws(spreadsheet):
    try:
        ws = spreadsheet.worksheet(WORKSHEET_NAME)
    except gspread.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title=WORKSHEET_NAME, rows=DEFAULT_SHEET_ROWS, cols=len(SHEET_COLUMNS))
        ws.append_row(SHEET_COLUMNS, value_input_option="USER_ENTERED")
        auto_format_absensi_sheet(ws)
        return ws

    if ws.row_count < DEFAULT_SHEET_ROWS:
        ws.resize(rows=DEFAULT_SHEET_ROWS)

    header = ws.row_values(1)
    if header != SHEET_COLUMNS:
        ws.resize(cols=max(ws.col_count, len(SHEET_COLUMNS)))
        ws.update("A1", [SHEET_COLUMNS], value_input_option="USER_ENTERED")
        auto_format_absensi_sheet(ws)
    return ws


# =========================
# DROPBOX
# =========================
@st.cache_resource
def connect_dropbox():
    if "dropbox" not in st.secrets or "access_token" not in st.secrets["dropbox"]:
        raise RuntimeError("Dropbox secrets tidak ditemukan: dropbox.access_token")

    dbx = dropbox.Dropbox(st.secrets["dropbox"]["access_token"])
    dbx.users_get_current_account()
    return dbx


def upload_selfie_to_dropbox(dbx, img_bytes: bytes, nama: str, ts_file: str, ext: str) -> Tuple[str, str]:
    clean_name = sanitize_name(nama).replace(" ", "_") or "Unknown"
    filename = f"{ts_file}_selfie{ext}"
    path = f"{DROPBOX_ROOT}/{clean_name}/{filename}"

    dbx.files_upload(img_bytes, path, mode=dropbox.files.WriteMode.add)

    settings = SharedLinkSettings(requested_visibility=RequestedVisibility.public)
    url = "-"
    try:
        link = dbx.sharing_create_shared_link_with_settings(path, settings=settings)
        url = link.url
    except ApiError as e:
        try:
            if e.error.is_shared_link_already_exists():
                links = dbx.sharing_list_shared_links(path, direct_only=True).links
                if links:
                    url = links[0].url
        except Exception:
            url = "-"

    url_raw = url.replace("?dl=0", "?raw=1") if url and url != "-" else "-"
    return url_raw, path


# =========================
# REKAP
# =========================
def normalize_posisi(text: str) -> str:
    t = str(text or "").strip().lower()
    t = t.replace("&", " dan ")
    t = re.sub(r"[/,_\\-\\.]+", " ", t)
    t = re.sub(r"[^a-z0-9\\s]+", " ", t)
    t = re.sub(r"\\s+", " ", t).strip()
    return t


POSISI_ALIASES: Dict[str, str] = {
    "spv": "supervisor",
    "sup": "supervisor",
    "super visor": "supervisor",
    "supervisior": "supervisor",
    "admin": "administrasi",
    "adm": "administrasi",
    "kry": "karyawan",
    "karyawan": "karyawan",
    "staf": "staff",
    "staff": "staff",
    "teknisi": "teknisi",
    "technician": "teknisi",
    "driver": "driver",
    "drv": "driver",
    "security": "security",
    "satpam": "security",
}


def smart_canonical_posisi(raw_pos: str, known_canon: List[str]) -> str:
    p = normalize_posisi(raw_pos)
    if not p:
        return ""
    if p in POSISI_ALIASES:
        p = POSISI_ALIASES[p]
    if known_canon:
        best = difflib.get_close_matches(p, known_canon, n=1, cutoff=0.88)
        if best:
            return best[0]
    return p


def display_posisi(canon: str) -> str:
    if not canon:
        return "-"
    return " ".join(w.capitalize() for w in canon.split())


def parse_date_prefix(ts: str) -> str:
    s = str(ts or "").strip()
    if not s:
        return ""
    try:
        dt = datetime.strptime(s, "%d-%m-%Y %H:%M:%S")
        return dt.strftime("%d-%m-%Y")
    except Exception:
        return s[:10]


def _group_contiguous_rows(rows: List[int]) -> List[Tuple[int, int]]:
    if not rows:
        return []
    rows = sorted(rows)
    ranges = []
    start = prev = rows[0]
    for r in rows[1:]:
        if r == prev + 1:
            prev = r
        else:
            ranges.append((start, prev))
            start = prev = r
    ranges.append((start, prev))
    return ranges


@st.cache_data(ttl=30, show_spinner=False)
def get_rekap_today() -> Dict:
    sh = connect_gsheet()
    ws = get_or_create_ws(sh)

    today_str = now_local().strftime("%d-%m-%Y")

    ts_col = ws.col_values(1)
    if not ts_col or len(ts_col) < 2:
        return {"today": today_str, "total": 0, "dup_removed": 0, "by_pos": [], "all_people": []}

    match_rows = []
    for idx, ts in enumerate(ts_col[1:], start=2):
        if parse_date_prefix(ts) == today_str:
            match_rows.append(idx)

    if not match_rows:
        return {"today": today_str, "total": 0, "dup_removed": 0, "by_pos": [], "all_people": []}

    ranges = _group_contiguous_rows(match_rows)
    data = []
    for a, b in ranges:
        chunk = ws.get(f"A{a}:D{b}")
        if chunk:
            data.extend(chunk)

    seen_keys = set()
    dup_removed = 0
    people_by_pos = defaultdict(list)
    all_people = []
    known_canon: List[str] = []

    for r in data:
        ts = (r[0] if len(r) > 0 else "") or ""
        nama = (r[1] if len(r) > 1 else "") or ""
        hp = (r[2] if len(r) > 2 else "") or ""
        pos = (r[3] if len(r) > 3 else "") or ""

        if parse_date_prefix(ts) != today_str:
            continue

        nama_clean = sanitize_name(nama)
        hp_clean = sanitize_phone(hp)
        key = hp_clean if hp_clean else nama_clean.lower().strip()
        if not key:
            continue

        if key in seen_keys:
            dup_removed += 1
            continue
        seen_keys.add(key)

        pos_canon = smart_canonical_posisi(pos, known_canon)
        if pos_canon and pos_canon not in known_canon:
            known_canon.append(pos_canon)

        who = nama_clean if nama_clean else (hp_clean if hp_clean else "Tanpa Nama")
        who_display = f"{who} ({hp_clean})" if hp_clean and who else who

        all_people.append({
            "Nama": who,
            "No HP/WA": hp_clean or "-",
            "Posisi": display_posisi(pos_canon) if pos_canon else "-",
            "Timestamp": ts,
        })
        people_by_pos[pos_canon if pos_canon else "(tanpa posisi)"].append(who_display)

    by_pos = []
    for canon, people in people_by_pos.items():
        by_pos.append({
            "Posisi": display_posisi(canon) if canon != "(tanpa posisi)" else "Tanpa Posisi",
            "Jumlah": len(people),
            "Yang Hadir": ", ".join(people),
        })
    by_pos.sort(key=lambda x: (-x["Jumlah"], x["Posisi"].lower()))

    return {
        "today": today_str,
        "total": len(seen_keys),
        "dup_removed": dup_removed,
        "by_pos": by_pos,
        "all_people": all_people,
    }


# =========================
# EXPORT
# =========================
HYPERLINK_RE = re.compile(r'=HYPERLINK\\("(?P<url>.*?)"\\s*,\\s*"(?P<label>.*?)"\\)', re.IGNORECASE)


def extract_hyperlink_url(formula_or_value: str) -> str:
    s = str(formula_or_value or "").strip()
    if not s:
        return ""
    m = HYPERLINK_RE.match(s)
    if not m:
        if s.startswith("http://") or s.startswith("https://"):
            return s
        return ""
    url = m.group("url").replace('""', '"')
    return url


def make_csv_bytes(header: List[str], rows: List[List[str]]) -> bytes:
    """
    CSV rapi untuk Excel Indonesia:
    - delimiter ';'
    - tambah baris pertama 'sep=;' agar Excel auto-parse
    - encoding utf-8-sig agar aman untuk Excel
    """
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    buf.write("sep=;\\n")
    writer.writerow(header)
    for r in rows:
        writer.writerow(r)
    return buf.getvalue().encode("utf-8-sig")


def make_xlsx_bytes(sheet_name: str, header: List[str], rows: List[List[str]], hyperlink_col: Optional[int] = None) -> bytes:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl belum terpasang. Tambahkan 'openpyxl' ke requirements.txt")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_fill = PatternFill("solid", fgColor="EAF3FF")
    header_font = Font(bold=True, color="0A2540")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(header)
    for col_idx in range(1, len(header) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align

    ws.freeze_panes = "A2"

    for r in rows:
        ws.append(r)

    body_align = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(header)):
        for cell in row:
            cell.alignment = body_align

    # hyperlink_col = index 0-based
    if hyperlink_col is not None and 0 <= hyperlink_col < len(header):
        col_excel = hyperlink_col + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_excel)
            url = str(cell.value or "").strip()
            if url.startswith("http://") or url.startswith("https://"):
                cell.value = "Bukti Foto"
                cell.hyperlink = url
                cell.font = Font(color="0B66E4", underline="single")

    # Column widths rapi
    preset_widths = {}
    for i, col_name in enumerate(header):
        name = col_name.lower()
        if "timestamp" in name:
            preset_widths[i + 1] = 20
        elif "nama" in name:
            preset_widths[i + 1] = 24
        elif "no hp" in name or "wa" in name:
            preset_widths[i + 1] = 16
        elif "posisi" in name:
            preset_widths[i + 1] = 18
        elif "dropbox" in name:
            preset_widths[i + 1] = 46
        elif "bukti" in name or "selfie" in name:
            preset_widths[i + 1] = 18
        else:
            preset_widths[i + 1] = 18

    for col_idx, w in preset_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_export_rekap_today(rekap: Dict) -> Tuple[List[str], List[List[str]]]:
    header = ["Timestamp", "Nama", "No HP/WA", "Posisi"]
    rows = []
    for p in rekap.get("all_people", []):
        rows.append([
            str(p.get("Timestamp", "")),
            str(p.get("Nama", "")),
            str(p.get("No HP/WA", "")),
            str(p.get("Posisi", "")),
        ])
    return header, rows


def fetch_log_full() -> Tuple[List[str], List[List[str]]]:
    sh = connect_gsheet()
    ws = get_or_create_ws(sh)

    try:
        values = ws.get("A:F", value_render_option="FORMULA")
    except TypeError:
        values = ws.get("A:F")

    if not values or len(values) < 2:
        return SHEET_COLUMNS, []

    data_rows = values[1:]

    rows = []
    for r in data_rows:
        r = (r + [""] * 6)[:6]
        ts, nama, hp, pos, bukti, dbx_path = r

        url = extract_hyperlink_url(bukti)
        bukti_out = url if url else ""

        if not (
            str(ts).strip()
            or str(nama).strip()
            or str(hp).strip()
            or str(pos).strip()
            or str(bukti_out).strip()
            or str(dbx_path).strip()
        ):
            continue

        rows.append([
            str(ts).strip(),
            str(nama).strip(),
            str(hp).strip(),
            str(pos).strip(),
            str(bukti_out).strip(),  # URL
            str(dbx_path).strip(),
        ])

    export_header = [COL_TIMESTAMP, COL_NAMA, COL_HP, COL_POSISI, "Bukti Selfie (URL)", COL_DBX_PATH]
    return export_header, rows


# =========================
# SESSION DEFAULTS
# =========================
if "saving" not in st.session_state:
    st.session_state.saving = False
if "submitted_once" not in st.session_state:
    st.session_state.submitted_once = False
if "selfie_method" not in st.session_state:
    st.session_state.selfie_method = "Upload"

if "export_ready" not in st.session_state:
    st.session_state.export_ready = False
if "export_xlsx" not in st.session_state:
    st.session_state.export_xlsx = None
if "export_csv" not in st.session_state:
    st.session_state.export_csv = None
if "export_base_name" not in st.session_state:
    st.session_state.export_base_name = ""


# =========================
# PAGES
# =========================
mode = get_mode()

# ===== PAGE: QR / ADMIN
if mode != "absen":
    render_header("QR Absensi", f"{BRAND_TAGLINE} • Scan untuk Absensi")

    st.markdown(
        """
<div class="jala-card">
  <div style="font-weight:700; font-size:16px; margin-bottom:6px;">QR Code Absensi</div>
  <div class="jala-muted">
    Tempel QR ini di area masuk. Karyawan scan QR → isi form → selfie → submit.
  </div>
</div>
        """,
        unsafe_allow_html=True,
    )
    st.write("")

    if not QR_URL:
        st.warning("QR URL belum diset. Isi `app.qr_url` di secrets.")
        st.code("Contoh: https://YOUR-APP.streamlit.app/?mode=absen", language="text")
        st.stop()

    if ENABLE_TOKEN and TOKEN_SECRET:
        if "token=" not in QR_URL:
            sep = "&" if "?" in QR_URL else "?"
            qr_url_effective = f"{QR_URL}{sep}token={TOKEN_SECRET}"
        else:
            qr_url_effective = QR_URL
    else:
        qr_url_effective = QR_URL

    qr_png = build_qr_png(qr_url_effective)

    st.markdown('<div class="jala-card">', unsafe_allow_html=True)
    st.image(qr_png, caption="QR Absensi", use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)

    st.write("")
    st.download_button(
        "⬇️ Download QR",
        data=qr_png,
        file_name="qr_absensi_jala.png",
        mime="image/png",
        use_container_width=True
    )

    with st.expander("ℹ️ Info Admin"):
        st.write("**Link Form Absensi:**")
        st.code(qr_url_effective, language="text")
        st.caption("Gunakan link ini untuk kebutuhan admin. Untuk karyawan, gunakan QR.")

    st.markdown(
        f"""
<div style="text-align:center; margin-top: 10px;" class="jala-muted">
  © {BRAND_TAGLINE} • Absensi QR
</div>
        """,
        unsafe_allow_html=True,
    )
    st.stop()


# ===== PAGE: ABSEN
dt = now_local()
ts_display = dt.strftime("%d-%m-%Y %H:%M:%S")
ts_file = dt.strftime("%Y-%m-%d_%H-%M-%S")

render_header("Form Absensi", f"{BRAND_TAGLINE} • {ts_display} ({TZ_NAME})")

if ENABLE_TOKEN and TOKEN_SECRET:
    incoming_token = get_token_from_url()
    if incoming_token != TOKEN_SECRET:
        st.error("Akses tidak valid. Silakan scan QR resmi dari kantor.")
        st.stop()

st.markdown(
    """
<div class="jala-card">
  <div style="font-weight:700; font-size:16px; margin-bottom:6px;">Petunjuk</div>
  <div class="jala-muted">
    Isi data karyawan, pilih metode selfie, lalu submit.
    Jika kamera bermasalah, gunakan <b>Upload</b>.
  </div>
</div>
    """,
    unsafe_allow_html=True,
)
st.write("")

with st.form("form_absen", clear_on_submit=False):
    st.subheader("1) Data Karyawan")
    nama = st.text_input("Nama Lengkap", placeholder="Contoh: Andi Saputra")
    no_hp = st.text_input("No HP/WA", placeholder="Contoh: 08xxxxxxxxxx atau +628xxxxxxxxxx")
    posisi = st.text_input("Posisi / Jabatan", placeholder="Contoh: Driver / Teknisi / Supervisor")

    st.markdown('<div class="jala-divider"></div>', unsafe_allow_html=True)

    st.subheader("2) Selfie Kehadiran")
    method = st.radio(
        "Metode selfie",
        options=["Upload (lebih stabil)", "Kamera (jika HP mendukung)"],
        index=0 if st.session_state.selfie_method == "Upload" else 1,
        horizontal=False,
    )
    st.session_state.selfie_method = "Upload" if method.startswith("Upload") else "Kamera"

    selfie_cam = None
    selfie_upload = None

    if st.session_state.selfie_method == "Kamera":
        st.caption("Jika kamera blank/lemot, pilih Upload.")
        selfie_cam = st.camera_input("Ambil selfie")
    else:
        st.caption("Foto akan dioptimalkan otomatis agar hemat kuota.")
        selfie_upload = st.file_uploader("Upload foto selfie", type=["jpg", "jpeg", "png"])

    st.markdown('<div class="jala-divider"></div>', unsafe_allow_html=True)

    submit = st.form_submit_button(
        "✅ Submit Absensi",
        disabled=st.session_state.saving or st.session_state.submitted_once,
        use_container_width=True,
    )

if submit:
    if st.session_state.submitted_once:
        st.warning("Absensi sudah tersimpan. Jika ingin absen lagi, refresh halaman.")
        st.stop()

    nama_clean = sanitize_name(nama)
    hp_clean = sanitize_phone(no_hp)
    posisi_final = str(posisi).strip()
    img_bytes, ext = get_selfie_bytes(selfie_cam, selfie_upload)

    errors = []
    if not nama_clean:
        errors.append("• Nama wajib diisi.")
    if not hp_clean or len(hp_clean.replace("+", "")) < 8:
        errors.append("• No HP/WA wajib diisi (minimal 8 digit).")
    if not posisi_final:
        errors.append("• Posisi wajib diisi.")
    if img_bytes is None:
        errors.append("• Selfie wajib (kamera atau upload).")

    if errors:
        st.error("Mohon lengkapi dulu:\n\n" + "\n".join(errors))
        st.stop()

    st.session_state.saving = True
    try:
        with st.spinner("Menyimpan absensi..."):
            img_bytes_opt, ext_opt = optimize_image_bytes(img_bytes, ext)

            sh = connect_gsheet()
            ws = get_or_create_ws(sh)
            dbx = connect_dropbox()

            link_selfie, dbx_path = upload_selfie_to_dropbox(dbx, img_bytes_opt, nama_clean, ts_file, ext_opt)
            link_cell = make_hyperlink(link_selfie, "Bukti Foto")

            ws.append_row(
                [ts_display, nama_clean, hp_clean, posisi_final, link_cell, dbx_path],
                value_input_option="USER_ENTERED"
            )

        get_rekap_today.clear()
        st.session_state.submitted_once = True
        st.success("Absensi berhasil tersimpan. Terima kasih ✅")

        if st.button("↩️ Isi ulang (reset form)", use_container_width=True):
            st.session_state.saving = False
            st.session_state.submitted_once = False
            st.session_state.selfie_method = "Upload"
            st.rerun()

    except AuthError:
        st.error("Dropbox token tidak valid. Hubungi admin.")
    except Exception as e:
        st.error("Gagal menyimpan absensi.")
        with st.expander("Detail error (untuk admin)"):
            st.code(str(e))
    finally:
        st.session_state.saving = False


# ===== REKAP UI
st.write("")
st.subheader("📊 Rekap Kehadiran (Hari ini)")

try:
    rekap = get_rekap_today()

    top1, top2 = st.columns([1, 1])
    with top1:
        st.metric("Total hadir", rekap["total"])
    with top2:
        if st.button("🔄 Refresh rekap", use_container_width=True):
            get_rekap_today.clear()
            st.rerun()

    st.caption(f"Tanggal: **{rekap['today']}**")

    if rekap["dup_removed"] > 0:
        st.info(
            f"Catatan: terdeteksi **{rekap['dup_removed']}** entri duplikat (No HP/Nama sama) "
            f"dan tidak dihitung agar rekap akurat."
        )

    if rekap["total"] == 0:
        st.warning("Belum ada absensi untuk hari ini.")
    else:
        st.markdown(
            """
<div class="jala-card" style="margin-bottom: 10px;">
  <div style="font-weight:700; margin-bottom:8px;">Klasifikasi jumlah hadir per posisi</div>
</div>
            """,
            unsafe_allow_html=True,
        )
        render_table(rekap["by_pos"], columns=["Posisi", "Jumlah", "Yang Hadir"], min_width_px=640)

        with st.expander("👥 Lihat siapa saja yang sudah datang (detail)"):
            render_table(rekap["all_people"], columns=["Nama", "No HP/WA", "Posisi", "Timestamp"], min_width_px=640)

    # ===== EXPORT
    with st.expander("⬇️ Download Rekap (Excel / CSV)"):
        st.markdown(
            """
<div class="jala-muted">
Unduh data dengan format rapi:
<b>XLSX</b> (paling aman untuk Excel) atau <b>CSV</b> (pakai delimiter <code>;</code> agar tidak jadi 1 kolom).
</div>
            """,
            unsafe_allow_html=True,
        )

        scope = st.radio(
            "Pilih data yang diunduh",
            options=["Rekap Hari Ini (dedup)", "Log Lengkap (semua data)"],
            index=0,
            horizontal=False,
        )

        cA, cB = st.columns([1, 1])
        with cA:
            prep = st.button("📦 Siapkan File", use_container_width=True)
        with cB:
            if st.button("🧹 Reset file", use_container_width=True):
                st.session_state.export_ready = False
                st.session_state.export_xlsx = None
                st.session_state.export_csv = None
                st.session_state.export_base_name = ""
                st.rerun()

        if prep:
            try:
                with st.spinner("Menyiapkan file export..."):
                    ts_tag = now_local().strftime("%Y-%m-%d_%H-%M")

                    if scope.startswith("Rekap"):
                        header, rows = build_export_rekap_today(rekap)
                        base = f"rekap_hadir_{rekap['today'].replace('-', '')}_{ts_tag}"
                        xlsx = make_xlsx_bytes("Rekap Hari Ini", header, rows, hyperlink_col=None)
                        csv_b = make_csv_bytes(header, rows)
                    else:
                        header, rows = fetch_log_full()
                        base = f"log_absensi_{ts_tag}"
                        # hyperlink_col = kolom "Bukti Selfie (URL)" => index 4
                        xlsx = make_xlsx_bytes("Log Absensi", header, rows, hyperlink_col=4)
                        csv_b = make_csv_bytes(header, rows)

                    st.session_state.export_xlsx = xlsx
                    st.session_state.export_csv = csv_b
                    st.session_state.export_base_name = base
                    st.session_state.export_ready = True

                st.success("File siap diunduh ✅")
            except Exception as ex:
                st.error("Gagal menyiapkan file export.")
                st.code(str(ex))

        if st.session_state.export_ready and st.session_state.export_xlsx and st.session_state.export_csv:
            st.write("")
            d1, d2 = st.columns(2)
            with d1:
                st.download_button(
                    "⬇️ Download XLSX",
                    data=st.session_state.export_xlsx,
                    file_name=f"{st.session_state.export_base_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with d2:
                st.download_button(
                    "⬇️ Download CSV",
                    data=st.session_state.export_csv,
                    file_name=f"{st.session_state.export_base_name}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )

except Exception as e:
    st.warning("Rekap kehadiran belum bisa ditampilkan (cek koneksi GSheet).")
    with st.expander("Detail error (untuk admin)"):
        st.code(str(e))

st.markdown(
    f"""
<div style="text-align:center; margin-top: 14px;" class="jala-muted">
  © {BRAND_TAGLINE} • Absensi QR
</div>
    """,
    unsafe_allow_html=True,
)
